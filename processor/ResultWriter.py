import os
import shutil
import subprocess
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy

SOFFICE_FALLBACK_PATHS = [
    '/Applications/LibreOffice.app/Contents/MacOS/soffice',
    r'C:\Program Files\LibreOffice\program\soffice.exe',
    '/usr/bin/soffice',
    '/usr/local/bin/soffice',
]


class ResultWriter:
    RACE_COUNT = 6
    FIRST_OUTPUT_WRITE_ROW = 8
    RESULT_COLUMN_COUNT = 3 + 2 + 4 * (RACE_COUNT - 1)

    def __init__(self, config, category_sum_results):
        self.config = config
        self.category_sum_results = category_sum_results
        self.outputFileName = "vysledky%s.xlsx" % self.config.year
        self.wb = None
        self.template_sheet = None

    def write(self):
        if os.path.exists(self.outputFileName):
            os.remove(self.outputFileName)

        self.wb = load_workbook('outputTemplate.xlsx')
        self.template_sheet = self.wb.get_sheet_by_name('Template')

        for cat_results in self.category_sum_results:
            self.prepare_sheet(cat_results.category.name, self.config.year, cat_results.category.get_title(), len(cat_results.personal_results))

            ws = self.wb.get_sheet_by_name(cat_results.category.name)
            row = self.FIRST_OUTPUT_WRITE_ROW
            for pr in cat_results.personal_results:
                ws.cell(row=row, column=2).value = pr.person.name
                ws.cell(row=row, column=3).value = pr.person.team
                ws.cell(row=row, column=4).value = pr.person.birth_year

                r = pr.race_results[0]
                self.write_position_and_points(ws.cell(row=row, column=5), ws.cell(row=row, column=6), r)

                for i in range(1, len(pr.race_results)):
                    r = pr.race_results[i]
                    self.write_position_and_points(ws.cell(row=row, column=7 + (i - 1) * 4),
                                                   ws.cell(row=row, column=8 + (i - 1) * 4), r)
                    if r.sum_points is not None:
                        ws.cell(row=row, column=9 + (i - 1) * 4).value = r.sum_position
                        ws.cell(row=row, column=10 + (i - 1) * 4).value = r.sum_points

                row = row + 1

        self.wb.remove_sheet(self.template_sheet)
        self.wb.save(self.outputFileName)
        self.wb.close()

        self.export_pdf()

    def export_pdf(self):
        soffice = self._find_soffice()
        if not soffice:
            print("WARNING: 'soffice' (LibreOffice) not found — skipping PDF export. "
                  "Set SOFFICE_BIN env var or install LibreOffice.", file=sys.stderr)
            return

        xlsx_path = os.path.abspath(self.outputFileName)
        out_dir = os.path.dirname(xlsx_path) or '.'
        intermediate_pdf = os.path.join(out_dir, "vysledky%s.pdf" % self.config.year)
        target_pdf = os.path.join(out_dir, "VKCT %s.pdf" % self.config.year)

        if os.path.exists(intermediate_pdf):
            os.remove(intermediate_pdf)

        user_install = "file://%s" % os.path.join(
            os.path.expanduser("~"), ".cache", "lo_pdfexport_vkct"
        )
        cmd = [
            soffice,
            "--headless",
            "-env:UserInstallation=%s" % user_install,
            "--convert-to", "pdf:calc_pdf_Export",
            "--outdir", out_dir,
            xlsx_path,
        ]

        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        except (subprocess.TimeoutExpired, OSError) as e:
            print("WARNING: PDF export failed: %s" % e, file=sys.stderr)
            return

        if result.returncode != 0 or not os.path.exists(intermediate_pdf):
            print("WARNING: PDF export failed (exit=%d):\n%s\n%s"
                  % (result.returncode, result.stdout, result.stderr), file=sys.stderr)
            return

        if os.path.exists(target_pdf):
            os.remove(target_pdf)
        os.rename(intermediate_pdf, target_pdf)
        print("PDF exported: %s" % target_pdf)

    @staticmethod
    def _find_soffice():
        env_bin = os.environ.get("SOFFICE_BIN")
        if env_bin and os.path.exists(env_bin):
            return env_bin
        which = shutil.which("soffice")
        if which:
            return which
        for path in SOFFICE_FALLBACK_PATHS:
            if os.path.exists(path):
                return path
        return None

    @staticmethod
    def write_position_and_points(pos_cell, points_cell, race_result):
        if race_result.position is not None:
            pos_cell.value = race_result.position if not race_result.half_points else "%s*" % race_result.position
            points_cell.value = race_result.points
            # highlight ignored results
            if race_result.ignored_in_summary:
                font = copy(points_cell.font)
                font.italic = True
                font.color = 'FF808080'
                points_cell.font = font

    def prepare_sheet(self, cat_name, year, cat_title, line_count):
        last_output_column = self.RESULT_COLUMN_COUNT + 1

        ws = self.wb.copy_worksheet(self.template_sheet)
        ws.title = cat_name
        ws.cell(row=2, column=2).value = "%s %d" % (ws.cell(row=2, column=2).value, year)
        ws.cell(row=3, column=2).value = cat_title
        for col in range(2, last_output_column + 1):
            src = ws.cell(row=self.FIRST_OUTPUT_WRITE_ROW, column=col)
            for row in range(self.FIRST_OUTPUT_WRITE_ROW + 1, self.FIRST_OUTPUT_WRITE_ROW + line_count):
                n = ws.cell(row=row, column=col)
                n.border = src.border.copy()
                n.fill = src.fill.copy()
                n.font = src.font.copy()
                n.alignment = src.alignment.copy()

        ws.freeze_panes = ws.cell(row=self.FIRST_OUTPUT_WRITE_ROW, column=5)
        ws.print_area = f'B2:{get_column_letter(last_output_column)}{line_count - 1 + self.FIRST_OUTPUT_WRITE_ROW:d}'
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False
        Worksheet.set_printer_settings(ws, paper_size=Worksheet.PAPERSIZE_A4, orientation=Worksheet.ORIENTATION_LANDSCAPE)
