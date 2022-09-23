# coding: utf-8
import difflib
import re
from collections import namedtuple

from openpyxl import load_workbook

from ResultWriter import ResultWriter


class Person:
    def __init__(self, name, team, birth_year):
        self.name = name
        self.team = team
        self.birth_year = birth_year

    def get_key(self):
        return ("%s@%s" % (self.name, self.birth_year)).encode('utf-8')


class PersonForCheck:
    def __init__(self, person: Person):
        self.person = person
        self.sourceYears = list[str]()


class ResultLine:
    def __init__(self, person, position, approved_pos, is_alternative):
        self.person = person
        self.position = position
        self.approved_pos = approved_pos
        self.is_alternative = is_alternative


class RaceResult:
    def __init__(self):
        self.position = None
        self.points = None
        self.sum_points = None
        self.sum_position = None
        self.half_points = False
        self.ignored_in_summary = False


class PersonalResults:
    def __init__(self, person, race_count):
        self.person = person
        self.race_results = list[RaceResult]()
        for i in range(race_count):
            self.race_results.append(RaceResult())


class CategorySummaryResults:
    def __init__(self, category):
        self.category = category
        self.personal_results = []


class Category:
    def __init__(self, name, min_age, max_age, count_positions, inputs, current_year):
        self.name = name
        self.min_age = min_age
        self.max_age = max_age
        self.count_positions = count_positions
        self.inputs = inputs
        self.results = list[list[ResultLine]]()
        self.min_year = current_year - max_age
        self.max_year = current_year - min_age

    def get_race_count(self):
        i = 0
        for ipt in self.inputs:
            if not ipt.is_alternative:
                i = i + 1
        return i

    def get_title(self):
        if self.min_age == 0:
            return u'%s (%s a mladší)' % (self.name, self.min_year)
        if self.max_age == 100:
            return u'%s (%s a starší)' % (self.name, self.max_year)
        return u'%s (%s - %s)' % (self.name, self.min_year, self.max_year)


CategoryInput = namedtuple("CategoryInput",
                           "file_name sheet_name, first_row, name_col, name2_col, team_col, birth_year_col, pos_col, is_alternative")
Config = namedtuple("Config", "year, max_race_count, categories")


class ParsePosition:
    def __init__(self):
        self.file = None
        self.sheet = None
        self.row = None


parsePosition = ParsePosition()


def main():
    process_results('config2022.xlsx')
    # build_people_list('configValidator.xlsx')
    pass


def build_people_list(validator_config_file):
    validation_config = load_config(validator_config_file)
    info('Loading old results...')
    read_results(validation_config, False)

    info('Building people dir...')
    people_dir = dict[str, PersonForCheck]()
    for category in validation_config.categories:
        sourceYears = build_input_year_list(category)

        for i in range(len(category.results)):
            resultList = category.results[i]
            for resLine in resultList:
                p = resLine.person
                existing = people_dir.get(p.get_key())
                if not existing:
                    existing = PersonForCheck(p)
                    people_dir[p.get_key()] = existing
                existing.sourceYears.append(sourceYears[i])

    info('Verifying people dir...')
    people_list = list(people_dir.values())
    for i in range(len(people_list)):
        pi = people_list[i]
        if not pi.person.birth_year:
            continue
        for j in range(i + 1, len(people_list)):
            pj = people_list[j]
            if not pj.person.birth_year or abs(pi.person.birth_year - pj.person.birth_year) > 5:
                # ignore too big differences
                continue
            ratio = get_names_matching_ratio(pi.person.name, pj.person.name)
            if ratio > 0.9:
                warning("Similar names in old results: '%s(%s) [%s]' ~ '%s(%s) [%s]' (%f)"
                        % (pi.person.name, pi.person.birth_year, ', '.join(pi.sourceYears),
                           pj.person.name, pj.person.birth_year, ', '.join(pj.sourceYears), ratio))


def build_input_year_list(category):
    inputYears = list[str]()
    for i in category.inputs:
        m = re.match(r'.*\\(\d{4})\\', i.file_name)
        if not m or not m.group(1):
            error("Failed to extract year from input file path: '%s'" % i.file_name)
        inputYears.append(m.group(1))
    return inputYears


def process_results(config_file):
    config = load_config(config_file)
    info('Loading results...')
    read_results(config, True)
    info('Filling missing birth years...')
    fill_missing_birth_years(config)
    info('Counting results...')
    category_sum_results = extract_summary_results(config)
    complete_summary_results(category_sum_results, config.max_race_count)
    info('Checking names...')
    check_names(category_sum_results)
    info('Writing output...')
    writer = ResultWriter(config, category_sum_results)
    writer.write()
    info('Done.')


def fill_missing_birth_years(config):
    for cat in config.categories:
        # build map of birth years
        birth_years = dict[str, Person]()
        for res in cat.results:
            for resLine in res:
                p = resLine.person
                if p.birth_year is not None and p.name not in birth_years:
                    birth_years[p.name] = p.birth_year

        # fill missing birth years
        for res in cat.results:
            for resLine in res:
                p = resLine.person
                if p.birth_year is None and p.name in birth_years:
                    p.birth_year = birth_years[p.name]


def check_names(category_sum_results):
    for cat_results in category_sum_results:
        for i in range(len(cat_results.personal_results)):
            for j in range(i + 1, len(cat_results.personal_results)):
                pi = cat_results.personal_results[i]
                pj = cat_results.personal_results[j]
                ratio = get_names_matching_ratio(pi.person.name, pj.person.name)
                if ratio > 0.8:
                    warning("Similar names in category '%s': '%s(%s) [%s]' ~ '%s(%s) [%s]' (%f)"
                            % (cat_results.category.name,
                               pi.person.name, pi.person.birth_year, get_race_index_list(pi),
                               pj.person.name, pj.person.birth_year, get_race_index_list(pj), ratio))
                pass
        pass


def get_names_matching_ratio(name_a, name_b):
    parts_a = name_a.split()

    res = 0
    if len(parts_a) < len(name_b.split()):
        return get_names_matching_ratio(name_b, name_a)
    if len(parts_a) == 3:
        res = max(
            difflib.SequenceMatcher(a="%s %s" % (parts_a[0], parts_a[2]), b=name_b).ratio(),
            difflib.SequenceMatcher(a="%s %s" % (parts_a[1], parts_a[2]), b=name_b).ratio()
        )
    return max(res, difflib.SequenceMatcher(a=name_a, b=name_b).ratio())


def get_race_index_list(personal_results):
    result = []
    for i in range(len(personal_results.race_results)):
        if personal_results.race_results[i].position is not None:
            result.append(str(i + 1))
    return ', '.join(result)


def read_results(config, validate_values: bool):
    for cat in config.categories:
        for i in cat.inputs:
            res = read_result_sheet(i.file_name, i.sheet_name, i.first_row, i.name_col, i.name2_col, i.team_col,
                                    i.birth_year_col, i.pos_col, i.is_alternative, cat, validate_values)
            cat.results.append(res)


def extract_summary_results(config):
    category_sum_results = []
    for cat in config.categories:
        cat_res = CategorySummaryResults(cat)
        res_map = {}
        category_sum_results.append(cat_res)
        race_idx = -1
        for i in range(0, len(cat.results)):
            race_results = cat.results[i]

            if not cat.inputs[i].is_alternative:
                race_idx = race_idx + 1

            for res_line in race_results:
                key = res_line.person.get_key()
                pr = res_map.get(key)
                if pr is None:
                    res_map[key] = pr = PersonalResults(res_line.person, cat.get_race_count())
                elif pr.person.team is None and res_line.person.team is not None:
                    # fix missing team if present
                    pr.person = Person(pr.person.name, res_line.person.team, pr.person.birth_year)

                rr = pr.race_results[race_idx]
                rr.position = res_line.position

                rr.points = compute_points(len(race_results), res_line.position, res_line.is_alternative,
                                           cat.count_positions)
                rr.half_points = res_line.is_alternative

        cat_res.personal_results = list(res_map.values())
    return category_sum_results


point_table = [50, 45, 40, 37, 34, 31, 28, 26, 24, 22, 20, 19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3,
               2, 1]


def compute_points(people_count, position, half_points, count_positions):
    if position in DNF_ACRONYMS:
        return 0

    # nejmensi kategorie jen bod za ucast
    if not count_positions:
        return 1

    if half_points:
        p = max(0, 26 - position)
    else:
        bonus = min(people_count, 30)
        p = max(bonus - (position - 1), 0)
        if position - 1 < len(point_table):
            p = p + point_table[position - 1]

    return p


def mark_ignored_results(race_results: list[RaceResult], max_race_count: int):
    sorted_results = sorted(race_results, key=lambda rr_: get_nullable_as_int(rr_.points), reverse=True)[max_race_count:]
    for rr in sorted_results:
        rr.ignored_in_summary = True


def sum_race_results(race_results: list[RaceResult], max_race_count: int):
    sorted_results = sorted(race_results, key=lambda rr_: get_nullable_as_int(rr_.points), reverse=True)[:max_race_count]
    res = None
    for rr in sorted_results:
        if rr.points is not None:
            res = rr.points if res is None else res + rr.points
    return res


def complete_summary_results(category_sum_results, max_race_count: int):
    for cat_results in category_sum_results:
        # compute sum_points
        if len(cat_results.personal_results) == 0:
            continue

        for pr in cat_results.personal_results:
            if cat_results.category.count_positions:
                mark_ignored_results(pr.race_results, max_race_count)

            for i in range(0, len(pr.race_results)):
                # don't limit max race count for categories without counting of positions
                max_ = max_race_count if cat_results.category.count_positions else i + 1
                pr.race_results[i].sum_points = sum_race_results(pr.race_results[:i + 1], max_)

        race_count = cat_results.category.get_race_count()
        # sort by first column if there is only one race
        if race_count == 1:
            cat_results.personal_results = sorted(cat_results.personal_results, key=lambda pr_: pr_.race_results[0].points, reverse=True)

        # complete sum_position
        for i in range(1, race_count):
            cat_results.personal_results = sorted(cat_results.personal_results, key=lambda pr_: get_nullable_as_int(pr_.race_results[i].sum_points), reverse=True)
            sum_pos = 1
            last_sum_points = None
            last_sum_position = None
            for pr in cat_results.personal_results:
                if pr.race_results[i].sum_points == last_sum_points:
                    pr.race_results[i].sum_position = last_sum_position
                else:
                    last_sum_points = pr.race_results[i].sum_points
                    pr.race_results[i].sum_position = last_sum_position = sum_pos
                sum_pos = sum_pos + 1


def get_nullable_as_int(sum_points):
    if sum_points is None:
        return -1
    return sum_points


def load_config(config_file):
    wb = load_workbook(config_file)
    ws = wb['Kategorie']

    current_year = ws['B1'].value
    max_race_count = ws['B2'].value
    categories = []
    row = 5
    while ws.cell(row=row, column=1).value is not None:
        cat_name = ws.cell(row=row, column=1).value
        category = Category(cat_name,
                            ws.cell(row=row, column=2).value,
                            ws.cell(row=row, column=3).value,
                            ws.cell(row=row, column=6).value == 1,
                            load_category_input_config(wb, cat_name),
                            current_year)
        categories.append(category)
        row = row + 1

    return Config(current_year, max_race_count, categories)


def load_category_input_config(wb, category_name):
    ws = wb[category_name]

    input_configs = []
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        file_name = ws.cell(row=row, column=1).value
        sheet_name = ws.cell(row=row, column=2).value
        first_row = ws.cell(row=row, column=3).value
        name_col = get_column_index(ws.cell(row=row, column=4).value)
        name2_col = get_column_index(ws.cell(row=row, column=5).value)
        team_col = get_column_index(ws.cell(row=row, column=6).value)
        birth_year_col = get_column_index(ws.cell(row=row, column=7).value)
        pos_col = get_column_index(ws.cell(row=row, column=8).value)
        is_alternative = ws.cell(row=row, column=9).value == 1

        input_configs.append(CategoryInput(
            file_name, sheet_name, first_row, name_col, name2_col, team_col, birth_year_col, pos_col, is_alternative
        ))
        row = row + 1
    return input_configs


def get_column_index(col_name):
    if not col_name:
        return None
    return ord(col_name[0]) - ord('A') + 1


def read_result_sheet(file_name, sheet_name, first_row, name_col, name2_col, team_col, birth_year_col, pos_col,
                      is_alternative, category, validate_values: bool):
    wb = load_workbook(file_name)
    try:
        ws = wb[sheet_name]
    except Exception:
        error("Failed to get sheet '%s' from %s" % (sheet_name, file_name))
        raise

    parsePosition.file = file_name
    parsePosition.sheet = sheet_name

    lines = []
    row = first_row
    while True:
        parsePosition.row = row
        name_val = ws.cell(row=row, column=name_col).value
        if not name_val or name_val.isspace():
            break

        if name2_col:
            name_val = name_val + ' ' + ws.cell(row=row, column=name2_col).value

        birth_year_cell = ws.cell(row=row, column=birth_year_col)
        position_cell = ws.cell(row=row, column=pos_col)
        line = create_normalized_result_line(
            name_val,
            ws.cell(row=row, column=team_col).value,
            birth_year_cell.value, has_approved_value(birth_year_cell),
            position_cell.value, has_approved_value(position_cell),
            is_alternative, category, validate_values)

        if line is not None:
            lines.append(line)
        row = row + 1

    parsePosition.file = None
    if validate_values:
        validate_positions(lines, sheet_name, file_name, is_alternative)
    return lines


def validate_positions(lines: list[ResultLine], sheet_name, file_name, is_alternative: bool):
    all_are_first = True
    for ln in lines:
        if not ln.position or ln.position in DNF_ACRONYMS:
            continue
        if ln.position and ln.position != 1:
            all_are_first = False
            break

    if all_are_first:
        # all are first, nothing to check
        return

    pos_dir = dict()
    # verify duplicates
    for ln in lines:
        if not ln.position or ln.position in DNF_ACRONYMS:
            continue
        if not ln.approved_pos and ln.position in pos_dir.keys():
            error("Non-unique position '%s' in sheet '%s' of %s. " % (ln.position, sheet_name, file_name))
        else:
            pos_dir[ln.position] = ln

    if not is_alternative:
        # check missing positions
        for i in range(1, len(pos_dir) + 1):
            if i not in pos_dir.keys():
                error("Missing position '%s' in sheet '%s' of %s. " % (i, sheet_name, file_name))


def has_approved_value(cell):
    return cell.font.b and cell.font.i and cell.font.u == 'single'


def create_normalized_result_line(name, team, birth_year, approved_birth_year, pos, approved_pos, is_alternative, category, validate_values: bool):
    n_name = normalize_name(name)

    birth_year = to_int(birth_year)
    if not isinstance(birth_year, int):
        warning("Birth year '%s' of %s is not a number" % (birth_year, name))
    elif birth_year == -1:
        birth_year = None
    elif validate_values and (birth_year < category.min_year or birth_year > category.max_year):
        if not approved_birth_year:
            if is_alternative:
                # info("Alternative result of '%s' (%d) is out of category age range %s (%d-%d). Skipping."
                #      % (name, birth_year, category.name, category.min_year, category.max_year))
                return None
            else:
                warning("Person '%s' (%d) is out of category age range %s (%d-%d). If you are sure, then make input field underlined, bold & italic."
                        % (name, birth_year, category.name, category.min_year, category.max_year))

    n_pos = to_int(pos)
    if not isinstance(n_pos, int) and n_pos not in DNF_ACRONYMS:
        info("Position '%s' is not a number! DNF '%s'!" % (n_pos, n_name))
        n_pos = 'DNF'

    return ResultLine(Person(n_name, team, birth_year), n_pos, approved_pos, is_alternative)


def to_int(n):
    if isinstance(n, int):
        return n
    if isinstance(n, str):
        if n.isdecimal():
            return int(n)
        if len(n) > 1 and n[-1] == '.' and n[:-1].isdecimal():
            return int(n[:-1])
    return n


def normalize_name(src):
    parts = src.replace('Mgr.', '').replace('Ing.', '').replace('ml.', '').replace('st.', '').strip().split()
    if len(parts) == 2:
        n1 = parts[0].capitalize()
        n2 = parts[1].capitalize()
        if n1 in first_names and n2 not in first_names:
            return "%s %s" % (n1, n2)
        if n2 in first_names and n1 not in first_names:
            return "%s %s" % (n2, n1)
        if "%s_%s" % (n1, n2) in first_names:
            return "%s %s" % (n1, n2)
        if "%s_%s" % (n2, n1) in first_names:
            return "%s %s" % (n2, n1)
        warning("Unable to detect first and second name: '%s'" % src)
        return src
    if len(parts) == 3:
        n1 = parts[0].capitalize()
        n2 = parts[1].capitalize()
        n3 = parts[2].capitalize()
        if n1 in first_names and n2 in first_names and n3 not in first_names:
            return "%s %s %s" % (n1, n2, n3)
        if n1 not in first_names and n2 in first_names and n3 in first_names:
            return "%s %s %s" % (n2, n3, n1)
        if "%s_%s_%s" % (n1, n2, n3) in first_names:
            return "%s %s %s" % (n1, n2, n3)
        if "%s_%s_%s" % (n2, n3, n1) in first_names:
            return "%s %s %s" % (n2, n3, n1)

    error("Unexpected name format, 2 parts expected: '%s'" % src)
    return src


def info(msg):
    message("INFO: ", msg)


def warning(msg):
    message("WARNING: ", msg)


def error(msg):
    message("ERROR: ", msg)


def message(prefix, msg):
    if parsePosition.file:
        print("%s%s @%s[%s]:%d" % (prefix, msg, parsePosition.file, parsePosition.sheet, parsePosition.row))
    else:
        print("%s%s" % (prefix, msg))


def load_first_names():
    res = {""}
    for l in open('firstNames.txt', encoding="utf8").read().split():
        res.add(l)
    return res


DNF_ACRONYMS = ['DNF', 'DNP', 'DNS']

first_names = load_first_names()

main()
